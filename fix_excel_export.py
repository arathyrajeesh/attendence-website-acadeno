"""
Fix: Replace export_batch_excel with a version that sets cell values
BEFORE calling merge_cells (required by newer openpyxl versions).
"""
import pathlib

views_path = pathlib.Path("attendence/students/views.py")
text = views_path.read_text(encoding="utf-8")

# Find and replace the entire export_batch_excel function
old_marker = "def export_batch_excel(request, pk):\n    \"\"\"\n    Horizontal-layout Excel export"
new_func = '''def export_batch_excel(request, pk):
    """
    Horizontal-layout Excel export matching the Students' Attendance Report format.
    SAFE version: sets cell values BEFORE merge_cells to avoid MergedCell read-only error.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    batch    = get_object_or_404(Batch, pk=pk)
    students = list(Student.objects.filter(batch=batch).order_by('name'))
    month, year = _get_month_year(request)

    records_qs = (
        Attendance.objects
        .filter(student__batch=batch)
        .select_related('student')
        .order_by('date')
    )
    if month and year:
        records_qs = records_qs.filter(date__year=year, date__month=month)
        period_label = f"{calendar.month_name[month]} {year}"
    else:
        period_label = "All Records"

    all_dates = sorted(set(r.date for r in records_qs))
    lookup = {}
    for r in records_qs:
        lookup.setdefault(r.student_id, {})[r.date] = r

    start_date = all_dates[0]  if all_dates else None
    end_date   = all_dates[-1] if all_dates else None

    DATES_PER_SHEET = 7
    chunks = [all_dates[i:i + DATES_PER_SHEET]
              for i in range(0, max(len(all_dates), 1), DATES_PER_SHEET)]
    if not all_dates:
        chunks = [[]]

    # ── Style helpers ─────────────────────────────────────────────────────────
    TB = Border(
        left=Side(style='thin', color='AAAAAA'),  right=Side(style='thin', color='AAAAAA'),
        top=Side(style='thin', color='AAAAAA'),   bottom=Side(style='thin', color='AAAAAA'),
    )
    C = Alignment(horizontal='center', vertical='center', wrap_text=True)
    L = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    HDR_FILL  = PatternFill('solid', fgColor='D9E1F2')
    SUB_FILL  = PatternFill('solid', fgColor='BDD7EE')
    DATE_FILL = PatternFill('solid', fgColor='DEEAF1')
    PRES_FILL = PatternFill('solid', fgColor='C6EFCE')
    ABS_FILL  = PatternFill('solid', fgColor='FFB3B3')
    NC_FILL   = PatternFill('solid', fgColor='FFF2CC')

    def sc(ws, row, col, value='', bold=False, size=9, color='000000',
           align=None, fill=None, border=None):
        """Set cell value+style BEFORE any merge."""
        cell = ws.cell(row=row, column=col, value=value)
        cell.font      = Font(bold=bold, size=size, color=color, name='Calibri')
        cell.alignment = align  or C
        if fill:   cell.fill   = fill
        if border: cell.border = border
        return cell

    def merge(ws, r1, c1, r2, c2):
        if (r1, c1) != (r2, c2):
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    wb = openpyxl.Workbook()

    for sh_idx, dates in enumerate(chunks):
        ws = wb.active if sh_idx == 0 else wb.create_sheet()
        ws.title = f"Sheet {sh_idx + 1}"

        n_dates    = len(dates)
        total_cols = 1 + n_dates * 2

        # ── ROW 1: Title ──────────────────────────────────────────────────────
        sc(ws, 1, 1, "STUDENTS\' ATTENDANCE REPORT",
           bold=True, size=14, align=C, fill=HDR_FILL, border=TB)
        merge(ws, 1, 1, 1, total_cols)
        ws.row_dimensions[1].height = 30

        # ── ROW 2: College / Internship / Duration ────────────────────────────
        sc(ws, 2, 1, "COLLEGE NAME :", bold=True, size=9, align=L, fill=HDR_FILL, border=TB)
        merge(ws, 2, 1, 2, 4)

        if total_cols >= 9:
            sc(ws, 2, 5, "INTERNSHIP :", bold=True, size=9, align=L, fill=HDR_FILL, border=TB)
            merge(ws, 2, 5, 2, 8)
            dur_col = 9
        else:
            dur_col = 5  # right after college block

        dur_text = f"DURATION: {batch.duration_display}"
        sc(ws, 2, dur_col, dur_text, bold=True, size=9, align=L, fill=HDR_FILL, border=TB)
        if dur_col <= total_cols:
            merge(ws, 2, dur_col, 2, total_cols)
        ws.row_dimensions[2].height = 18

        # ── ROW 3: Values ─────────────────────────────────────────────────────
        sc(ws, 3, 1, batch.name, size=9, align=L, fill=HDR_FILL, border=TB)
        merge(ws, 3, 1, 3, 4)

        if total_cols >= 9:
            sc(ws, 3, 5, batch.course, size=9, align=L, fill=HDR_FILL, border=TB)
            merge(ws, 3, 5, 3, 8)

        if start_date and end_date:
            date_range = (
                f"START DATE: {start_date.strftime('%d %b %Y').upper()}     "
                f"END DATE: {end_date.strftime('%d %b %Y').upper()}"
            )
        else:
            date_range = ""
        sc(ws, 3, dur_col, date_range, size=9, align=L, fill=HDR_FILL, border=TB)
        if dur_col <= total_cols:
            merge(ws, 3, dur_col, 3, total_cols)
        ws.row_dimensions[3].height = 18

        # ── ROW 4: Date headers ───────────────────────────────────────────────
        sc(ws, 4, 1, '', bold=True, size=9, fill=DATE_FILL, border=TB)
        for i, d in enumerate(dates):
            col = 2 + i * 2
            sc(ws, 4, col,   d.strftime('%d/%m/%Y'), bold=True, size=9,
               align=C, fill=DATE_FILL, border=TB)
            sc(ws, 4, col+1, '',                     bold=True, size=9,
               align=C, fill=DATE_FILL, border=TB)
            merge(ws, 4, col, 4, col + 1)
        ws.row_dimensions[4].height = 20

        # ── ROW 5: Sub-headers ────────────────────────────────────────────────
        sc(ws, 5, 1, 'NAME', bold=True, size=8, fill=SUB_FILL, border=TB)
        for i in range(n_dates):
            sc(ws, 5, 2 + i*2,   'TIME',       bold=True, size=8, fill=SUB_FILL, border=TB)
            sc(ws, 5, 3 + i*2,   'Attendance', bold=True, size=8, fill=SUB_FILL, border=TB)
        ws.row_dimensions[5].height = 18

        # ── DATA ROWS ─────────────────────────────────────────────────────────
        for s_idx, student in enumerate(students):
            row  = 6 + s_idx
            srec = lookup.get(student.id, {})

            sc(ws, row, 1, student.name, size=8, color='CC0000', align=L, border=TB)

            for i, d in enumerate(dates):
                rec    = srec.get(d)
                col_t  = 2 + i * 2
                col_a  = col_t + 1

                if rec:
                    if rec.login_time and rec.logout_time:
                        tv = (f"{rec.login_time.strftime('%I:%M%p').lstrip('0')}"
                              f"-{rec.logout_time.strftime('%I:%M%p').lstrip('0')}")
                    elif rec.login_time:
                        tv = rec.login_time.strftime('%I:%M%p').lstrip('0')
                    else:
                        tv = ''
                    av   = rec.get_status_display()
                    afill = (PRES_FILL if rec.status == 'present'
                             else ABS_FILL if rec.status == 'absent' else NC_FILL)
                else:
                    tv = av = ''
                    afill = None

                sc(ws, row, col_t, tv, size=8, border=TB)
                sc(ws, row, col_a, av, size=8, bold=bool(av), fill=afill, border=TB)

            ws.row_dimensions[row].height = 18

        # Empty grid rows
        for extra in range(5):
            r = 6 + len(students) + extra
            for c in range(1, total_cols + 1):
                sc(ws, r, c, '', size=8, border=TB)
            ws.row_dimensions[r].height = 15

        # ── FOOTER (last sheet only) ──────────────────────────────────────────
        if sh_idx == len(chunks) - 1:
            foot = 6 + len(students) + 7
            half = max(2, total_cols // 2)
            ver  = half + 1

            sc(ws, foot, 1, 'Notes', bold=True, size=9, align=L)
            merge(ws, foot, 1, foot, half)

            notes = [
                "1. This attendance sheet is prepared based on daily attendance records maintained by the organization.",
                "2. The details mentioned above are true and correct to the best of our knowledge.",
                "3. This document is issued only for submission to the concerned college/university.",
                "4. Attendance has been verified and approved by the undersigned authority.",
            ]
            for n_i, note in enumerate(notes):
                nr   = foot + 1 + n_i
                cell = ws.cell(row=nr, column=1, value=note)
                cell.font      = Font(size=8, color='0070C0', name='Calibri')
                cell.alignment = L
                merge(ws, nr, 1, nr, half)
                ws.row_dimensions[nr].height = 14

            sc(ws, foot, ver, 'Verification', bold=True, size=9, align=L)
            merge(ws, foot, ver, foot, total_cols)

            sig_r = foot + 2
            sc(ws, sig_r, ver,
               'Signature of Mentor/Supervisor: _______________________',
               size=8, align=L)
            merge(ws, sig_r, ver, sig_r, total_cols)

            name_r = foot + 4
            sc(ws, name_r, ver,
               'Name & Designation: _______________ / _______________',
               size=8, align=L)
            merge(ws, name_r, ver, name_r, total_cols)

        # ── Column widths ─────────────────────────────────────────────────────
        ws.column_dimensions['A'].width = 22
        for i in range(n_dates):
            ws.column_dimensions[get_column_letter(2 + i*2)].width = 15
            ws.column_dimensions[get_column_letter(3 + i*2)].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    period_slug = period_label.replace(' ', '_')
    fname = f"Batch_{batch.name.replace(' ', '_')}_{period_slug}_Attendance.xlsx"
    response = HttpResponse(
        buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f\'attachment; filename="{fname}"\\'
    return response

'''

if old_marker in text:
    # Find where export_batch_excel starts
    start = text.find(old_marker)
    # Find the next top-level def after it (export_batch_pdf)
    next_def = text.find('\ndef export_batch_pdf', start + 10)
    if next_def == -1:
        # It's the last function, cut to end
        text = text[:start] + new_func
    else:
        text = text[:start] + new_func + '\n' + text[next_def+1:]
    views_path.write_text(text, encoding='utf-8')
    print("✅ Patched export_batch_excel successfully")
else:
    print("❌ Marker not found — checking current function names...")
    import re
    fns = re.findall(r'^def (\w+)', text, re.MULTILINE)
    print("Functions found:", fns)
