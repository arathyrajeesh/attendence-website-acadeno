"""
Patch script: replaces everything from '# ── Batch Export Views' to EOF
in students/views.py with the new horizontal-layout batch export functions.
"""
import pathlib

views_path = pathlib.Path("attendence/students/views.py")
text = views_path.read_bytes().decode("utf-8", errors="replace")

# Keep everything before the batch export section
marker = "# \u2500\u2500 Batch Export Views"
cut = text.find(marker)
if cut == -1:
    # Try ASCII fallback
    marker = "# -- Batch Export Views"
    cut = text.find(marker)
if cut == -1:
    raise RuntimeError("Could not find batch export marker in views.py")

header = text[:cut]

new_section = r'''# ── Batch Export Views ─────────────────────────────────────────────────────────

def export_batch_excel(request, pk):
    """
    Horizontal-layout Excel export matching the Students' Attendance Report format:
    Dates as columns, students as rows, TIME + Attendance sub-columns per date.
    Splits into multiple sheets (7 dates each). Last sheet includes Notes/Verification footer.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    batch    = get_object_or_404(Batch, pk=pk)
    students = list(Student.objects.filter(batch=batch).order_by('name'))
    month, year = _get_month_year(request)

    records_qs = (
        Attendance.objects
        .filter(student__batch=batch)
        .select_related('student')
        .order_by('date')
    )
    if month and year:
        records_qs = records_qs.filter(date__year=year, date__month=month)
        period_label = f"{calendar.month_name[month]} {year}"
    else:
        period_label = "All Records"

    all_dates = sorted(set(r.date for r in records_qs))

    # Build lookup: student_id -> {date -> record}
    lookup = {}
    for r in records_qs:
        lookup.setdefault(r.student_id, {})[r.date] = r

    start_date = all_dates[0]  if all_dates else None
    end_date   = all_dates[-1] if all_dates else None

    DATES_PER_SHEET = 7
    chunks = [all_dates[i:i+DATES_PER_SHEET] for i in range(0, max(len(all_dates), 1), DATES_PER_SHEET)]
    if not all_dates:
        chunks = [[]]

    # ── Style helpers ────────────────────────────────────────────────────────────
    def thin_border(sides='all'):
        s = Side(style='thin', color='AAAAAA')
        n = Side(style=None)
        if sides == 'all':
            return Border(left=s, right=s, top=s, bottom=s)
        return Border(left=s if 'l' in sides else n,
                      right=s if 'r' in sides else n,
                      top=s if 't' in sides else n,
                      bottom=s if 'b' in sides else n)

    C  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    L  = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    HDR_FILL  = PatternFill('solid', fgColor='D9E1F2')   # blue-grey header
    SUB_FILL  = PatternFill('solid', fgColor='BDD7EE')   # lighter sub-header
    DATE_FILL = PatternFill('solid', fgColor='DEEAF1')   # date row
    PRES_FILL = PatternFill('solid', fgColor='C6EFCE')   # present – green
    ABS_FILL  = PatternFill('solid', fgColor='FFB3B3')   # absent  – red
    NC_FILL   = PatternFill('solid', fgColor='FFF2CC')   # no-class – yellow

    tb = thin_border()

    wb = openpyxl.Workbook()

    for sh_idx, dates in enumerate(chunks):
        ws = wb.active if sh_idx == 0 else wb.create_sheet()
        ws.title = f"Sheet {sh_idx+1}"

        n_date_cols = len(dates)
        total_cols  = 1 + n_date_cols * 2     # NAME + (TIME+Att) * n_dates
        last_col_letter = get_column_letter(total_cols)

        def mc(r1, c1, r2, c2):
            if (r1, c1) != (r2, c2):
                ws.merge_cells(start_row=r1, start_column=c1,
                               end_row=r2,   end_column=c2)
            return ws.cell(r1, c1)

        # ── ROW 1 : Title ───────────────────────────────────────────────────────
        c = mc(1, 1, 1, total_cols)
        c.value     = "STUDENTS' ATTENDANCE REPORT"
        c.font      = Font(bold=True, size=16, name='Calibri')
        c.alignment = C
        c.fill      = HDR_FILL
        ws.row_dimensions[1].height = 32

        # ── ROWS 2-3 : Header info ───────────────────────────────────────────────
        info_mid  = max(5, total_cols // 2 - 1)
        info_end  = total_cols

        # College name
        mc(2, 1, 2, 4).value = "COLLEGE NAME :"
        ws.cell(2, 1).font = Font(bold=True, size=9, name='Calibri')
        ws.cell(2, 1).alignment = L

        mc(3, 1, 3, 4).value = batch.name
        ws.cell(3, 1).font      = Font(size=9, name='Calibri')
        ws.cell(3, 1).alignment = L

        # Internship / course
        if total_cols >= 8:
            mc(2, 5, 2, 8).value = "INTERNSHIP :"
            ws.cell(2, 5).font = Font(bold=True, size=9, name='Calibri')
            ws.cell(2, 5).alignment = L

            mc(3, 5, 3, 8).value = batch.course
            ws.cell(3, 5).font      = Font(size=9, name='Calibri')
            ws.cell(3, 5).alignment = L

        # Duration / dates (right side)
        dur_col = min(9, total_cols)
        if dur_col <= total_cols:
            mc(2, dur_col, 2, total_cols).value = f"DURATION: {batch.duration_display}"
            ws.cell(2, dur_col).font = Font(bold=True, size=9, name='Calibri')
            ws.cell(2, dur_col).alignment = L

            if start_date and end_date:
                date_range = (
                    f"START DATE: {start_date.strftime('%d %b %Y').upper()}     "
                    f"END DATE: {end_date.strftime('%d %b %Y').upper()}"
                )
            else:
                date_range = ""
            mc(3, dur_col, 3, total_cols).value = date_range
            ws.cell(3, dur_col).font      = Font(size=9, name='Calibri')
            ws.cell(3, dur_col).alignment = L

        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 18

        # Outer border for header block
        for r in range(1, 4):
            for c_idx in range(1, total_cols + 1):
                cell = ws.cell(r, c_idx)
                cell.border = tb

        # ── ROW 4 : Date headers (merged pairs) ─────────────────────────────────
        c = ws.cell(4, 1, "")
        c.border = tb
        c.fill   = DATE_FILL

        for i, d in enumerate(dates):
            col = 2 + i * 2
            # Merge TIME + Attendance columns for the date label
            mc(4, col, 4, col + 1).value = d.strftime('%d/%m/%Y')
            ws.cell(4, col).font      = Font(bold=True, size=9, name='Calibri')
            ws.cell(4, col).alignment = C
            ws.cell(4, col).fill      = DATE_FILL
            ws.cell(4, col).border    = tb
            ws.cell(4, col + 1).border = tb
            ws.cell(4, col + 1).fill   = DATE_FILL

        ws.row_dimensions[4].height = 20

        # ── ROW 5 : Sub-headers NAME | TIME | Attendance … ──────────────────────
        for col in range(1, total_cols + 1):
            cell = ws.cell(5, col)
            cell.fill   = SUB_FILL
            cell.border = tb
            cell.font   = Font(bold=True, size=8, name='Calibri')
            cell.alignment = C

        ws.cell(5, 1).value = "NAME"
        for i in range(n_date_cols):
            ws.cell(5, 2 + i * 2).value     = "TIME"
            ws.cell(5, 3 + i * 2).value = "Attendance"

        ws.row_dimensions[5].height = 18

        # ── DATA ROWS ────────────────────────────────────────────────────────────
        for s_idx, student in enumerate(students):
            row  = 6 + s_idx
            srec = lookup.get(student.id, {})

            name_c = ws.cell(row, 1, student.name)
            name_c.font      = Font(size=8, color='CC0000', name='Calibri')
            name_c.alignment = L
            name_c.border    = tb

            for i, d in enumerate(dates):
                col_t = 2 + i * 2
                col_a = col_t + 1
                rec   = srec.get(d)

                if rec:
                    if rec.login_time and rec.logout_time:
                        time_val = (
                            f"{rec.login_time.strftime('%I:%M%p').lstrip('0')}"
                            f"-{rec.logout_time.strftime('%I:%M%p').lstrip('0')}"
                        )
                    elif rec.login_time:
                        time_val = rec.login_time.strftime('%I:%M%p').lstrip('0')
                    else:
                        time_val = ""
                    att_val  = rec.get_status_display()
                    att_fill = (PRES_FILL if rec.status == 'present'
                                else ABS_FILL if rec.status == 'absent'
                                else NC_FILL)
                else:
                    time_val = att_val = ""
                    att_fill = None

                tc = ws.cell(row, col_t, time_val)
                tc.font = Font(size=8, name='Calibri')
                tc.alignment = C
                tc.border    = tb

                ac = ws.cell(row, col_a, att_val)
                ac.font      = Font(size=8, bold=bool(att_val), name='Calibri')
                ac.alignment = C
                ac.border    = tb
                if att_fill:
                    ac.fill = att_fill

            ws.row_dimensions[row].height = 18

        # Empty grid rows below students
        for extra in range(5):
            r = 6 + len(students) + extra
            for c_idx in range(1, total_cols + 1):
                ws.cell(r, c_idx).border = tb
            ws.row_dimensions[r].height = 15

        # ── FOOTER (last sheet only) ─────────────────────────────────────────────
        if sh_idx == len(chunks) - 1:
            foot_row = 6 + len(students) + 7
            half_col = max(2, total_cols // 2)

            # Notes heading
            mc(foot_row, 1, foot_row, half_col).value = "Notes"
            ws.cell(foot_row, 1).font = Font(bold=True, size=9, name='Calibri')

            notes = [
                "1. This attendance sheet is prepared based on daily attendance records maintained by the organization.",
                "2. The details mentioned above are true and correct to the best of our knowledge.",
                "3. This document is issued only for submission to the concerned college/university.",
                "4. Attendance has been verified and approved by the undersigned authority.",
            ]
            for n_i, note in enumerate(notes):
                nr = foot_row + 1 + n_i
                mc(nr, 1, nr, half_col).value = note
                ws.cell(nr, 1).font      = Font(size=8, color='0070C0', name='Calibri')
                ws.cell(nr, 1).alignment = L
                ws.row_dimensions[nr].height = 14

            # Verification heading
            ver_col = half_col + 1
            mc(foot_row, ver_col, foot_row, total_cols).value = "Verification"
            ws.cell(foot_row, ver_col).font = Font(bold=True, size=9, name='Calibri')

            sig_row = foot_row + 2
            mc(sig_row, ver_col, sig_row, total_cols).value = (
                "Signature of Mentor/Supervisor: _______________________"
            )
            ws.cell(sig_row, ver_col).font      = Font(size=8, name='Calibri')
            ws.cell(sig_row, ver_col).alignment = L

            name_row = foot_row + 4
            mc(name_row, ver_col, name_row, total_cols).value = (
                "Name & Designation: _______________ / _______________"
            )
            ws.cell(name_row, ver_col).font      = Font(size=8, name='Calibri')
            ws.cell(name_row, ver_col).alignment = L

        # ── Column widths ────────────────────────────────────────────────────────
        ws.column_dimensions['A'].width = 22
        for i in range(n_date_cols):
            ws.column_dimensions[get_column_letter(2 + i * 2)].width = 15  # TIME
            ws.column_dimensions[get_column_letter(3 + i * 2)].width = 12  # Attendance

    # ── Save & return ────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    period_slug = period_label.replace(' ', '_')
    fname = f"Batch_{batch.name.replace(' ', '_')}_{period_slug}_Attendance.xlsx"
    response = HttpResponse(
        buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


def export_batch_pdf(request, pk):
    """
    Horizontal-layout PDF export (landscape A4) matching the Students' Attendance Report format.
    5 dates per page. Last page includes Notes/Verification footer.
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    batch    = get_object_or_404(Batch, pk=pk)
    students = list(Student.objects.filter(batch=batch).order_by('name'))
    month, year = _get_month_year(request)

    records_qs = (
        Attendance.objects
        .filter(student__batch=batch)
        .select_related('student')
        .order_by('date')
    )
    if month and year:
        records_qs = records_qs.filter(date__year=year, date__month=month)
        period_label = f"{calendar.month_name[month]} {year}"
    else:
        period_label = "All Records"

    all_dates = sorted(set(r.date for r in records_qs))

    lookup = {}
    for r in records_qs:
        lookup.setdefault(r.student_id, {})[r.date] = r

    start_date = all_dates[0]  if all_dates else None
    end_date   = all_dates[-1] if all_dates else None

    DATES_PER_PAGE = 5
    chunks = [all_dates[i:i+DATES_PER_PAGE]
              for i in range(0, max(len(all_dates), 1), DATES_PER_PAGE)]
    if not all_dates:
        chunks = [[]]

    # Colours
    C_HDR   = colors.HexColor('#D9E1F2')
    C_SUB   = colors.HexColor('#BDD7EE')
    C_DATE  = colors.HexColor('#DEEAF1')
    C_PRES  = colors.HexColor('#C6EFCE')
    C_ABS   = colors.HexColor('#FFB3B3')
    C_NC    = colors.HexColor('#FFF2CC')
    C_WHITE = colors.white
    C_DARK  = colors.HexColor('#1A1A1D')
    C_BLUE  = colors.HexColor('#0070C0')
    C_RED   = colors.HexColor('#CC0000')

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        topMargin=12*mm, bottomMargin=12*mm,
        leftMargin=12*mm, rightMargin=12*mm,
    )
    styles = getSampleStyleSheet()
    note_style = ParagraphStyle(
        'Note', parent=styles['Normal'], fontSize=7, textColor=C_BLUE, leading=10,
    )
    ver_style = ParagraphStyle(
        'Ver', parent=styles['Normal'], fontSize=7, leading=12,
    )

    story = []

    for ch_idx, dates in enumerate(chunks):
        n_dates  = len(dates)
        # Column widths: NAME=40mm, TIME=22mm, Att=18mm per date
        name_w   = 40 * mm
        time_w   = 22 * mm
        att_w    = 18 * mm
        col_widths = [name_w] + [time_w, att_w] * n_dates

        def para(text, bold=False, size=9, color=colors.black, align='CENTER'):
            sty = ParagraphStyle(
                'x', parent=styles['Normal'],
                fontSize=size, textColor=color, leading=size * 1.2,
                alignment={'LEFT': 0, 'CENTER': 1, 'RIGHT': 2}[align],
                fontName='Helvetica-Bold' if bold else 'Helvetica',
            )
            return Paragraph(text or '', sty)

        page_data = []

        # ── Title row ────────────────────────────────────────────────────────────
        title_span = 1 + n_dates * 2
        row_title = [para("STUDENTS' ATTENDANCE REPORT", bold=True, size=13)]
        row_title += [''] * (title_span - 1)
        page_data.append(row_title)

        # ── Header info rows ─────────────────────────────────────────────────────
        dur_text = batch.duration_display
        if start_date and end_date:
            dates_text = (
                f"START DATE: {start_date.strftime('%d %b %Y').upper()}   "
                f"END DATE: {end_date.strftime('%d %b %Y').upper()}"
            )
        else:
            dates_text = ""

        row_h1 = [para(f"<b>COLLEGE NAME :</b>  {batch.name}", size=8, align='LEFT')]
        row_h1 += [''] * (n_dates)
        row_h1 += [para(f"<b>INTERNSHIP :</b>  {batch.course}", size=8, align='LEFT')]
        row_h1 += [''] * (n_dates - 1)
        page_data.append(row_h1[:title_span])

        row_h2 = [para(f"<b>DURATION :</b>  {dur_text}", size=8, align='LEFT')]
        row_h2 += [''] * (n_dates)
        row_h2 += [para(dates_text, size=8, align='LEFT')]
        row_h2 += [''] * (n_dates - 1)
        page_data.append(row_h2[:title_span])

        # ── Date header row ───────────────────────────────────────────────────────
        row_dates = [para('', size=8)]
        for d in dates:
            row_dates.append(para(d.strftime('%d/%m/%Y'), bold=True, size=8))
            row_dates.append('')
        page_data.append(row_dates)

        # ── Sub-header row ────────────────────────────────────────────────────────
        row_sub = [para('NAME', bold=True, size=8)]
        for _ in dates:
            row_sub.append(para('TIME', bold=True, size=8))
            row_sub.append(para('Attendance', bold=True, size=8))
        page_data.append(row_sub)

        # ── Student data rows ─────────────────────────────────────────────────────
        for student in students:
            srec = lookup.get(student.id, {})
            row  = [para(student.name, size=8, color=C_RED, align='LEFT')]
            for d in dates:
                rec = srec.get(d)
                if rec:
                    if rec.login_time and rec.logout_time:
                        tv = (f"{rec.login_time.strftime('%I:%M%p').lstrip('0')}"
                              f"-{rec.logout_time.strftime('%I:%M%p').lstrip('0')}")
                    elif rec.login_time:
                        tv = rec.login_time.strftime('%I:%M%p').lstrip('0')
                    else:
                        tv = ''
                    av = rec.get_status_display()
                else:
                    tv = av = ''
                row.append(para(tv, size=7))
                row.append(para(av, size=7, bold=bool(av)))
            page_data.append(row)

        # Empty filler rows
        for _ in range(4):
            page_data.append([''] + [''] * (n_dates * 2))

        # ── Build table ───────────────────────────────────────────────────────────
        tbl = Table(page_data, colWidths=col_widths, repeatRows=0)

        # Merge spans: title, header rows, date cells
        n_stu = len(students)
        span_cmds = [
            # Title spans all columns
            ('SPAN', (0, 0), (title_span - 1, 0)),
            # Header info: each half spans half
            ('SPAN', (0, 1), (n_dates, 1)),
            ('SPAN', (n_dates + 1, 1), (title_span - 1, 1)),
            ('SPAN', (0, 2), (n_dates, 2)),
            ('SPAN', (n_dates + 1, 2), (title_span - 1, 2)),
            # Date merged pairs
        ]
        for i in range(n_dates):
            col = 1 + i * 2
            span_cmds.append(('SPAN', (col, 3), (col + 1, 3)))

        style_cmds = span_cmds + [
            # Outer grid
            ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#AAAAAA')),
            # Title
            ('BACKGROUND', (0, 0), (-1, 0), C_HDR),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, 0), 13),
            ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
            # Header info rows
            ('BACKGROUND', (0, 1), (-1, 2), C_HDR),
            ('FONTSIZE',   (0, 1), (-1, 2), 8),
            ('ALIGN',      (0, 1), (-1, 2), 'LEFT'),
            # Date row
            ('BACKGROUND', (0, 3), (-1, 3), C_DATE),
            ('FONTNAME',   (0, 3), (-1, 3), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 3), (-1, 3), 8),
            ('ALIGN',      (0, 3), (-1, 3), 'CENTER'),
            # Sub-header
            ('BACKGROUND', (0, 4), (-1, 4), C_SUB),
            ('FONTNAME',   (0, 4), (-1, 4), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 4), (-1, 4), 8),
            ('ALIGN',      (0, 4), (-1, 4), 'CENTER'),
            # Data rows
            ('FONTSIZE',   (0, 5), (-1, -1), 7),
            ('ALIGN',      (0, 5), (-1, -1), 'CENTER'),
            ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 5), (-1, 4 + n_stu),
             [colors.white, colors.HexColor('#F5F5F7')]),
            # Row heights
            ('ROWHEIGHT',  (0, 0), (0, 0), 24),
            ('ROWHEIGHT',  (0, 1), (0, 2), 15),
            ('ROWHEIGHT',  (0, 3), (0, 4), 16),
        ]

        # Colour attendance cells
        for s_i, student in enumerate(students):
            srec = lookup.get(student.id, {})
            data_row = 5 + s_i
            for d_i, d in enumerate(dates):
                rec = srec.get(d)
                if rec:
                    col_a = 2 + d_i * 2
                    fill  = (C_PRES if rec.status == 'present'
                             else C_ABS if rec.status == 'absent'
                             else C_NC)
                    style_cmds.append(('BACKGROUND', (col_a, data_row), (col_a, data_row), fill))

        tbl.setStyle(TableStyle(style_cmds))
        story.append(tbl)

        # ── Footer (last chunk only) ──────────────────────────────────────────────
        if ch_idx == len(chunks) - 1:
            story.append(Spacer(1, 8*mm))
            story.append(HRFlowable(width='100%', thickness=0.5,
                                    color=colors.HexColor('#AAAAAA')))
            story.append(Spacer(1, 4*mm))

            notes = [
                "1. This attendance sheet is prepared based on daily attendance records maintained by the organization.",
                "2. The details mentioned above are true and correct to the best of our knowledge.",
                "3. This document is issued only for submission to the concerned college/university.",
                "4. Attendance has been verified and approved by the undersigned authority.",
            ]
            footer_data = [
                [Paragraph('<b>Notes</b>', styles['Normal']),
                 Paragraph('<b>Verification</b>', styles['Normal'])],
                ['', ''],
            ]
            for i, note in enumerate(notes):
                left_cell = Paragraph(note, note_style)
                right_cell = ''
                if i == 1:
                    right_cell = Paragraph(
                        'Signature of Mentor/Supervisor: _______________________', ver_style)
                elif i == 3:
                    right_cell = Paragraph(
                        'Name &amp; Designation: _______________ / _______________', ver_style)
                footer_data.append([left_cell, right_cell])

            half = sum(col_widths) / 2
            ft = Table(footer_data, colWidths=[half, half])
            ft.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
            ]))
            story.append(ft)

        if ch_idx < len(chunks) - 1:
            from reportlab.platypus import PageBreak
            story.append(PageBreak())

    doc.build(story)
    buf.seek(0)

    period_slug = period_label.replace(' ', '_')
    fname = f"Batch_{batch.name.replace(' ', '_')}_{period_slug}_Attendance.pdf"
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response
'''

# Write patched file
out = header + new_section
views_path.write_text(out, encoding='utf-8')
print(f"Patched {views_path} ({len(out)} chars, marker at char {cut})")
