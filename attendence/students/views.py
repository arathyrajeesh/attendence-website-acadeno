import io
import calendar
import datetime
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, get_object_or_404
from django.urls import reverse_lazy
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from django.views.generic import UpdateView, DeleteView, DetailView
from django.views.generic.list import ListView
from django.views.generic.edit import FormMixin, CreateView
from django.db import transaction

from .models import Student, Attendance, Batch
from .forms import StudentForm, AttendanceForm, BatchForm


# ── Student Views ──────────────────────────────────────────────────────────────

class StudentListView(FormMixin, ListView):
    """Student list + inline add form on the same page."""
    model = Student
    template_name = 'students/student_list.html'
    context_object_name = 'students'
    ordering = ['-id']
    form_class = StudentForm
    success_url = reverse_lazy('student-list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['months'] = list(enumerate(calendar.month_name))[1:]
        context['batches'] = Batch.objects.order_by('-created_at')
        return context

    def post(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        form = self.get_form()
        if form.is_valid():
            form.save()
            messages.success(request, "Student added successfully.")
            return redirect(self.success_url)
        return self.render_to_response(self.get_context_data(form=form))


class StudentAttendanceDetailView(ListView):
    """All attendance records for a single student, with optional month/year filter."""
    model = Attendance
    template_name = 'students/student_attendance.html'
    context_object_name = 'attendances'

    def dispatch(self, request, *args, **kwargs):
        self.student = get_object_or_404(Student, pk=kwargs['pk'])
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        qs = Attendance.objects.filter(student=self.student).order_by('date')
        month = self.request.GET.get('month', '')
        year  = self.request.GET.get('year',  '')
        if month and year:
            try:
                qs = qs.filter(date__month=int(month), date__year=int(year))
            except ValueError:
                pass
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        attendances = context['attendances']
        present  = sum(1 for a in attendances if a.status == 'present')
        absent   = sum(1 for a in attendances if a.status == 'absent')
        no_class = sum(1 for a in attendances if a.status == 'no_class')
        total_class_days = present + absent
        pct = f"{(present / total_class_days * 100):.1f}%" if total_class_days else "N/A"

        context.update({
            'student':        self.student,
            'months':         list(enumerate(calendar.month_name))[1:],
            'selected_month': self.request.GET.get('month', ''),
            'selected_year':  self.request.GET.get('year',  ''),
            'present_count':  present,
            'absent_count':   absent,
            'no_class_count': no_class,
            'total_count':    len(attendances),
            'attendance_pct': pct,
        })
        return context


class StudentUpdateView(SuccessMessageMixin, UpdateView):
    model = Student
    form_class = StudentForm
    template_name = 'students/student_form.html'
    success_url = reverse_lazy('student-list')
    success_message = "Student %(name)s updated successfully."


class StudentDeleteView(DeleteView):
    model = Student
    template_name = 'students/student_confirm_delete.html'
    success_url = reverse_lazy('student-list')

    def post(self, request, *args, **kwargs):
        messages.success(request, "Student deleted successfully.")
        return super().post(request, *args, **kwargs)


# ── Batch Views ────────────────────────────────────────────────────────────────

class BatchImportView(FormMixin, ListView):
    """
    Batch/group import: enter a college name, course, hours+days duration,
    and paste a list of student names to add them all at once.
    """
    model = Batch
    template_name = 'students/batch_import.html'
    context_object_name = 'batches'
    ordering = ['-created_at']
    form_class = BatchForm
    success_url = reverse_lazy('batch-import')

    def post(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        form = BatchForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                batch = form.save()
                names = form.cleaned_data['parsed_names']
                students = [
                    Student(
                        name=n,
                        course=batch.course,
                        course_type=batch.course_type,
                        duration_hours=batch.duration_hours,
                        duration_days=batch.duration_days,
                        duration_months=batch.duration_months,
                        batch=batch,
                    )
                    for n in names
                ]
                Student.objects.bulk_create(students)
            messages.success(
                request,
                f"✅ Batch '{batch.name}' created with {len(names)} students "
                f"— {batch.get_course_type_display()} · {batch.duration_display}."
            )
            return redirect(self.success_url)
        return self.render_to_response(self.get_context_data(form=form))


class BatchDetailView(ListView):
    """Show all students belonging to a specific batch."""
    model = Student
    template_name = 'students/batch_detail.html'
    context_object_name = 'students'

    def dispatch(self, request, *args, **kwargs):
        self.batch = get_object_or_404(Batch, pk=kwargs['pk'])
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        return Student.objects.filter(batch=self.batch).order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['batch'] = self.batch
        return context


class BatchDeleteView(DeleteView):
    model = Batch
    template_name = 'students/batch_confirm_delete.html'
    success_url = reverse_lazy('batch-import')

    def post(self, request, *args, **kwargs):
        messages.success(request, "Batch deleted successfully.")
        return super().post(request, *args, **kwargs)


# ── Attendance Views ───────────────────────────────────────────────────────────

class AttendanceListView(FormMixin, ListView):
    """Attendance records + inline mark-attendance form on the same page."""
    model = Attendance
    template_name = 'students/attendance_list.html'
    context_object_name = 'attendances'
    form_class = AttendanceForm
    success_url = reverse_lazy('attendance-list')

    def get_queryset(self):
        date_str = self.request.GET.get('date', '')
        try:
            filter_date = datetime.date.fromisoformat(date_str) if date_str else datetime.date.today()
        except ValueError:
            filter_date = datetime.date.today()
        return (Attendance.objects
                .select_related('student')
                .filter(date=filter_date)
                .order_by('-id'))

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        date_str = self.request.GET.get('date', '')
        today = str(datetime.date.today())
        context['selected_date'] = date_str or today
        context['is_today'] = not date_str or date_str == today
        return context

    def post(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        form = self.get_form()
        if form.is_valid():
            form.save()
            messages.success(request, "Attendance marked successfully.")
            return redirect(self.success_url)
        return self.render_to_response(self.get_context_data(form=form))


class AttendanceUpdateView(SuccessMessageMixin, UpdateView):
    model = Attendance
    form_class = AttendanceForm
    template_name = 'students/attendance_form.html'
    success_url = reverse_lazy('attendance-list')
    success_message = "Attendance record updated successfully."


class AttendanceDeleteView(DeleteView):
    model = Attendance
    template_name = 'students/attendance_confirm_delete.html'
    success_url = reverse_lazy('attendance-list')

    def post(self, request, *args, **kwargs):
        messages.success(request, "Attendance record deleted successfully.")
        return super().post(request, *args, **kwargs)


# ── Export Helpers ─────────────────────────────────────────────────────────────

def _get_month_year(request):
    """Return (month, year) from GET params, or (None, None) for all-records export."""
    raw_month = request.GET.get('month', '').strip()
    raw_year  = request.GET.get('year',  '').strip()
    if not raw_month and not raw_year:
        return None, None
    today = datetime.date.today()
    try:
        month = int(raw_month) if raw_month else today.month
        year  = int(raw_year)  if raw_year  else today.year
        if not (1 <= month <= 12):
            month = today.month
    except (ValueError, TypeError):
        month, year = today.month, today.year
    return month, year


def export_student_pdf(request, pk):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    student = get_object_or_404(Student, pk=pk)
    month, year = _get_month_year(request)

    records = Attendance.objects.filter(student=student).order_by('date')
    if month and year:
        records = records.filter(date__year=year, date__month=month)
        period_label = f"{calendar.month_name[month]} {year}"
    else:
        period_label = "All Records"

    buf  = io.BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=A4,
                              topMargin=18*mm, bottomMargin=18*mm,
                              leftMargin=18*mm, rightMargin=18*mm)
    styles = getSampleStyleSheet()
    elems  = []

    elems.append(Paragraph("ATTENDANCE REPORT", styles['Title']))
    elems.append(Paragraph(f"{student.name}  —  {period_label}", styles['Heading2']))
    elems.append(Paragraph(
        f"Course: {student.course}  |  Duration: {student.duration_display}"
        + (f"  |  Batch: {student.batch.name}" if student.batch else ""),
        styles['Normal']
    ))
    elems.append(Spacer(1, 8*mm))

    data = [['Date', 'Mode', 'Login', 'Logout', 'Status']]
    present = absent = no_class = 0
    for r in records:
        if r.status == 'present':
            present += 1
        elif r.status == 'absent':
            absent += 1
        else:
            no_class += 1
        data.append([
            str(r.date),
            r.get_mode_display(),
            r.login_time.strftime('%H:%M') if r.login_time else '—',
            r.logout_time.strftime('%H:%M') if r.logout_time else '—',
            r.get_status_display(),
        ])

    if len(data) > 1:
        t = Table(data, colWidths=[38*mm, 30*mm, 24*mm, 24*mm, 24*mm])
        t.setStyle(TableStyle([
            ('BACKGROUND',   (0, 0), (-1, 0), colors.HexColor('#1a1a1d')),
            ('TEXTCOLOR',    (0, 0), (-1, 0), colors.white),
            ('FONTNAME',     (0, 0), (-1,  0), 'Helvetica-Bold'),
            ('FONTSIZE',     (0, 0), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f7')]),
            ('GRID',         (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
            ('ALIGN',        (0, 0), (-1, -1), 'CENTER'),
            ('PADDING',      (0, 0), (-1, -1), 5),
        ]))
        elems.append(t)
    else:
        elems.append(Paragraph("No attendance records for this period.", styles['Normal']))

    elems.append(Spacer(1, 8*mm))
    total_class_days = present + absent
    pct = f"{(present / total_class_days * 100):.1f}%" if total_class_days else "N/A"

    summary = Table(
        [['Present', 'Absent', 'No Class', 'Attendance %'],
         [str(present), str(absent), str(no_class), pct]],
        colWidths=[30*mm, 30*mm, 30*mm, 35*mm]
    )
    summary.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5b9cf6')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 9),
        ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
        ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
        ('PADDING',    (0, 0), (-1, -1), 6),
    ]))
    elems.append(Paragraph("Summary", styles['Heading3']))
    elems.append(Spacer(1, 3*mm))
    elems.append(summary)

    doc.build(elems)
    buf.seek(0)

    period_slug = period_label.replace(' ', '_')
    fname = f"{student.name.replace(' ', '_')}_{period_slug}.pdf"
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


def export_student_excel(request, pk):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    student = get_object_or_404(Student, pk=pk)
    month, year = _get_month_year(request)

    records = Attendance.objects.filter(student=student).order_by('date')
    if month and year:
        records = records.filter(date__year=year, date__month=month)
        period_label = f"{calendar.month_name[month]} {year}"
        tab_title    = f"{calendar.month_name[month][:3]} {year}"
    else:
        period_label = "All Records"
        tab_title    = "All Records"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tab_title

    thin = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC'),
    )
    center = Alignment(horizontal='center', vertical='center')

    duration_str = student.duration_display
    batch_str = f"  |  Batch: {student.batch.name}" if student.batch else ""
    for row, text in enumerate([
        "ATTENDANCE REPORT",
        f"{student.name}  —  {period_label}",
        f"Course: {student.course}  |  Duration: {duration_str}{batch_str}",
    ], start=1):
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = text
        ws[f'A{row}'].alignment = center
        ws[f'A{row}'].font = Font(bold=(row == 1), size=(13 if row == 1 else 10))

    ws.append([])  # blank

    headers = ['Date', 'Mode', 'Login', 'Logout', 'Status']
    ws.append(headers)
    hdr_fill = PatternFill('solid', fgColor='1A1A1D')
    for col in range(1, 6):
        c = ws.cell(row=5, column=col)
        c.fill = hdr_fill
        c.font = Font(color='FFFFFF', bold=True, size=9)
        c.alignment = center
        c.border = thin

    present = absent = no_class = 0
    for i, r in enumerate(records, start=6):
        if r.status == 'present':
            present += 1
        elif r.status == 'absent':
            absent += 1
        else:
            no_class += 1

        ws.append([
            str(r.date), r.get_mode_display(),
            r.login_time.strftime('%H:%M') if r.login_time else '—',
            r.logout_time.strftime('%H:%M') if r.logout_time else '—',
            r.get_status_display(),
        ])
        bg = 'FFFFFF' if i % 2 == 0 else 'F5F5F7'
        row_fill = PatternFill('solid', fgColor=bg)
        for col in range(1, 6):
            c = ws.cell(row=i, column=col)
            c.fill = row_fill; c.border = thin; c.alignment = center
        sc = ws.cell(row=i, column=5)
        status_color = '22C55E' if r.status == 'present' else ('EF4444' if r.status == 'absent' else 'F59E0B')
        sc.font = Font(color=status_color, bold=True, size=9)

    total_class_days = present + absent
    pct = f"{(present / total_class_days * 100):.1f}%" if total_class_days else "N/A"

    ws.append([])
    sr = ws.max_row + 1
    for col, val in enumerate(['Present', 'Absent', 'No Class', 'Attendance %', ''], start=1):
        c = ws.cell(row=sr, column=col, value=val)
        if col < 5:
            c.fill = PatternFill('solid', fgColor='5B9CF6')
            c.font = Font(color='FFFFFF', bold=True, size=9)
            c.alignment = center; c.border = thin
    for col, val in enumerate([present, absent, no_class, pct, ''], start=1):
        c = ws.cell(row=sr + 1, column=col, value=val)
        if col < 5:
            c.font = Font(bold=True, size=9)
            c.alignment = center; c.border = thin

    for col, w in zip('ABCDE', [14, 12, 10, 10, 10]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    period_slug = period_label.replace(' ', '_')
    fname = f"{student.name.replace(' ', '_')}_{period_slug}.xlsx"
    response = HttpResponse(buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response

# ── Batch Export Views ─────────────────────────────────────────────────────────

def export_batch_excel(request, pk):
    """
    Horizontal-layout Excel export matching the Students' Attendance Report format.
    SAFE version: sets cell values BEFORE merge_cells to avoid MergedCell read-only error.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    batch    = get_object_or_404(Batch, pk=pk)
    students = list(Student.objects.filter(batch=batch).order_by('name'))
    month, year = _get_month_year(request)

    records_qs = (
        Attendance.objects
        .filter(student__batch=batch)
        .select_related('student')
        .order_by('date')
    )
    if month and year:
        records_qs = records_qs.filter(date__year=year, date__month=month)
        period_label = f"{calendar.month_name[month]} {year}"
    else:
        period_label = "All Records"

    all_dates = sorted(set(r.date for r in records_qs))
    lookup = {}
    for r in records_qs:
        lookup.setdefault(r.student_id, {})[r.date] = r

    start_date = all_dates[0]  if all_dates else None
    end_date   = all_dates[-1] if all_dates else None

    DATES_PER_SHEET = 7
    chunks = [all_dates[i:i + DATES_PER_SHEET]
              for i in range(0, max(len(all_dates), 1), DATES_PER_SHEET)]
    if not all_dates:
        chunks = [[]]

    # ── Style helpers ─────────────────────────────────────────────────────────
    TB = Border(
        left=Side(style='thin', color='AAAAAA'),  right=Side(style='thin', color='AAAAAA'),
        top=Side(style='thin', color='AAAAAA'),   bottom=Side(style='thin', color='AAAAAA'),
    )
    C = Alignment(horizontal='center', vertical='center', wrap_text=True)
    L = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    HDR_FILL  = PatternFill('solid', fgColor='D9E1F2')
    SUB_FILL  = PatternFill('solid', fgColor='BDD7EE')
    DATE_FILL = PatternFill('solid', fgColor='DEEAF1')
    PRES_FILL = PatternFill('solid', fgColor='C6EFCE')
    ABS_FILL  = PatternFill('solid', fgColor='FFB3B3')
    NC_FILL   = PatternFill('solid', fgColor='FFF2CC')

    def sc(ws, row, col, value='', bold=False, size=9, color='000000',
           align=None, fill=None, border=None):
        """Set cell value+style BEFORE any merge."""
        cell = ws.cell(row=row, column=col, value=value)
        cell.font      = Font(bold=bold, size=size, color=color, name='Calibri')
        cell.alignment = align  or C
        if fill:   cell.fill   = fill
        if border: cell.border = border
        return cell

    def merge(ws, r1, c1, r2, c2):
        if (r1, c1) != (r2, c2):
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    wb = openpyxl.Workbook()

    for sh_idx, dates in enumerate(chunks):
        ws = wb.active if sh_idx == 0 else wb.create_sheet()
        ws.title = f"Sheet {sh_idx + 1}"

        n_dates    = len(dates)
        total_cols = 1 + n_dates * 2

        # ── ROW 1: Title ──────────────────────────────────────────────────────
        sc(ws, 1, 1, "STUDENTS' ATTENDANCE REPORT",
           bold=True, size=14, align=C, fill=HDR_FILL, border=TB)
        merge(ws, 1, 1, 1, total_cols)
        ws.row_dimensions[1].height = 30

        # ── ROW 2: College / Internship / Duration ────────────────────────────
        sc(ws, 2, 1, "COLLEGE NAME :", bold=True, size=9, align=L, fill=HDR_FILL, border=TB)
        merge(ws, 2, 1, 2, 4)

        if total_cols >= 9:
            sc(ws, 2, 5, "INTERNSHIP :", bold=True, size=9, align=L, fill=HDR_FILL, border=TB)
            merge(ws, 2, 5, 2, 8)
            dur_col = 9
        else:
            dur_col = 5  # right after college block

        dur_text = f"DURATION: {batch.duration_display}"
        sc(ws, 2, dur_col, dur_text, bold=True, size=9, align=L, fill=HDR_FILL, border=TB)
        if dur_col <= total_cols:
            merge(ws, 2, dur_col, 2, total_cols)
        ws.row_dimensions[2].height = 18

        # ── ROW 3: Values ─────────────────────────────────────────────────────
        sc(ws, 3, 1, batch.name, size=9, align=L, fill=HDR_FILL, border=TB)
        merge(ws, 3, 1, 3, 4)

        if total_cols >= 9:
            sc(ws, 3, 5, batch.course, size=9, align=L, fill=HDR_FILL, border=TB)
            merge(ws, 3, 5, 3, 8)

        if start_date and end_date:
            date_range = (
                f"START DATE: {start_date.strftime('%d %b %Y').upper()}     "
                f"END DATE: {end_date.strftime('%d %b %Y').upper()}"
            )
        else:
            date_range = ""
        sc(ws, 3, dur_col, date_range, size=9, align=L, fill=HDR_FILL, border=TB)
        if dur_col <= total_cols:
            merge(ws, 3, dur_col, 3, total_cols)
        ws.row_dimensions[3].height = 18

        # ── ROW 4: Date headers ───────────────────────────────────────────────
        sc(ws, 4, 1, '', bold=True, size=9, fill=DATE_FILL, border=TB)
        for i, d in enumerate(dates):
            col = 2 + i * 2
            sc(ws, 4, col,   d.strftime('%d/%m/%Y'), bold=True, size=9,
               align=C, fill=DATE_FILL, border=TB)
            sc(ws, 4, col+1, '',                     bold=True, size=9,
               align=C, fill=DATE_FILL, border=TB)
            merge(ws, 4, col, 4, col + 1)
        ws.row_dimensions[4].height = 20

        # ── ROW 5: Sub-headers ────────────────────────────────────────────────
        sc(ws, 5, 1, 'NAME', bold=True, size=8, fill=SUB_FILL, border=TB)
        for i in range(n_dates):
            sc(ws, 5, 2 + i*2,   'TIME',       bold=True, size=8, fill=SUB_FILL, border=TB)
            sc(ws, 5, 3 + i*2,   'Attendance', bold=True, size=8, fill=SUB_FILL, border=TB)
        ws.row_dimensions[5].height = 18

        # ── DATA ROWS ─────────────────────────────────────────────────────────
        for s_idx, student in enumerate(students):
            row  = 6 + s_idx
            srec = lookup.get(student.id, {})

            sc(ws, row, 1, student.name, size=8, color='CC0000', align=L, border=TB)

            for i, d in enumerate(dates):
                rec    = srec.get(d)
                col_t  = 2 + i * 2
                col_a  = col_t + 1

                if rec:
                    if rec.login_time and rec.logout_time:
                        tv = (f"{rec.login_time.strftime('%I:%M%p').lstrip('0')}"
                              f"-{rec.logout_time.strftime('%I:%M%p').lstrip('0')}")
                    elif rec.login_time:
                        tv = rec.login_time.strftime('%I:%M%p').lstrip('0')
                    else:
                        tv = ''
                    av   = rec.get_status_display()
                    afill = (PRES_FILL if rec.status == 'present'
                             else ABS_FILL if rec.status == 'absent' else NC_FILL)
                else:
                    tv = av = ''
                    afill = None

                sc(ws, row, col_t, tv, size=8, border=TB)
                sc(ws, row, col_a, av, size=8, bold=bool(av), fill=afill, border=TB)

            ws.row_dimensions[row].height = 18

        # Empty grid rows
        for extra in range(5):
            r = 6 + len(students) + extra
            for c in range(1, total_cols + 1):
                sc(ws, r, c, '', size=8, border=TB)
            ws.row_dimensions[r].height = 15

        # ── FOOTER (last sheet only) ──────────────────────────────────────────
        if sh_idx == len(chunks) - 1:
            foot = 6 + len(students) + 7
            half = max(2, total_cols // 2)
            ver  = half + 1

            sc(ws, foot, 1, 'Notes', bold=True, size=9, align=L)
            merge(ws, foot, 1, foot, half)

            notes = [
                "1. This attendance sheet is prepared based on daily attendance records maintained by the organization.",
                "2. The details mentioned above are true and correct to the best of our knowledge.",
                "3. This document is issued only for submission to the concerned college/university.",
                "4. Attendance has been verified and approved by the undersigned authority.",
            ]
            for n_i, note in enumerate(notes):
                nr   = foot + 1 + n_i
                cell = ws.cell(row=nr, column=1, value=note)
                cell.font      = Font(size=8, color='0070C0', name='Calibri')
                cell.alignment = L
                merge(ws, nr, 1, nr, half)
                ws.row_dimensions[nr].height = 14

            sc(ws, foot, ver, 'Verification', bold=True, size=9, align=L)
            merge(ws, foot, ver, foot, total_cols)

            sig_r = foot + 2
            sc(ws, sig_r, ver,
               'Signature of Mentor/Supervisor: _______________________',
               size=8, align=L)
            merge(ws, sig_r, ver, sig_r, total_cols)

            name_r = foot + 4
            sc(ws, name_r, ver,
               'Name & Designation: _______________ / _______________',
               size=8, align=L)
            merge(ws, name_r, ver, name_r, total_cols)

        # ── Column widths ─────────────────────────────────────────────────────
        ws.column_dimensions['A'].width = 22
        for i in range(n_dates):
            ws.column_dimensions[get_column_letter(2 + i*2)].width = 15
            ws.column_dimensions[get_column_letter(3 + i*2)].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    period_slug = period_label.replace(' ', '_')
    fname = f"Batch_{batch.name.replace(' ', '_')}_{period_slug}_Attendance.xlsx"
    response = HttpResponse(
        buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


def export_batch_pdf(request, pk):
    """
    Horizontal-layout PDF export (landscape A4) matching the Students' Attendance Report format.
    5 dates per page. Last page includes Notes/Verification footer.
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    batch    = get_object_or_404(Batch, pk=pk)
    students = list(Student.objects.filter(batch=batch).order_by('name'))
    month, year = _get_month_year(request)

    records_qs = (
        Attendance.objects
        .filter(student__batch=batch)
        .select_related('student')
        .order_by('date')
    )
    if month and year:
        records_qs = records_qs.filter(date__year=year, date__month=month)
        period_label = f"{calendar.month_name[month]} {year}"
    else:
        period_label = "All Records"

    all_dates = sorted(set(r.date for r in records_qs))

    lookup = {}
    for r in records_qs:
        lookup.setdefault(r.student_id, {})[r.date] = r

    start_date = all_dates[0]  if all_dates else None
    end_date   = all_dates[-1] if all_dates else None

    DATES_PER_PAGE = 5
    chunks = [all_dates[i:i+DATES_PER_PAGE]
              for i in range(0, max(len(all_dates), 1), DATES_PER_PAGE)]
    if not all_dates:
        chunks = [[]]

    # Colours
    C_HDR   = colors.HexColor('#D9E1F2')
    C_SUB   = colors.HexColor('#BDD7EE')
    C_DATE  = colors.HexColor('#DEEAF1')
    C_PRES  = colors.HexColor('#C6EFCE')
    C_ABS   = colors.HexColor('#FFB3B3')
    C_NC    = colors.HexColor('#FFF2CC')
    C_WHITE = colors.white
    C_DARK  = colors.HexColor('#1A1A1D')
    C_BLUE  = colors.HexColor('#0070C0')
    C_RED   = colors.HexColor('#CC0000')

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        topMargin=12*mm, bottomMargin=12*mm,
        leftMargin=12*mm, rightMargin=12*mm,
    )
    styles = getSampleStyleSheet()
    note_style = ParagraphStyle(
        'Note', parent=styles['Normal'], fontSize=7, textColor=C_BLUE, leading=10,
    )
    ver_style = ParagraphStyle(
        'Ver', parent=styles['Normal'], fontSize=7, leading=12,
    )

    story = []

    for ch_idx, dates in enumerate(chunks):
        n_dates  = len(dates)
        # Column widths: NAME=40mm, TIME=22mm, Att=18mm per date
        name_w   = 40 * mm
        time_w   = 22 * mm
        att_w    = 18 * mm
        col_widths = [name_w] + [time_w, att_w] * n_dates

        def para(text, bold=False, size=9, color=colors.black, align='CENTER'):
            sty = ParagraphStyle(
                'x', parent=styles['Normal'],
                fontSize=size, textColor=color, leading=size * 1.2,
                alignment={'LEFT': 0, 'CENTER': 1, 'RIGHT': 2}[align],
                fontName='Helvetica-Bold' if bold else 'Helvetica',
            )
            return Paragraph(text or '', sty)

        page_data = []

        # ── Title row ────────────────────────────────────────────────────────────
        title_span = 1 + n_dates * 2
        row_title = [para("STUDENTS' ATTENDANCE REPORT", bold=True, size=13)]
        row_title += [''] * (title_span - 1)
        page_data.append(row_title)

        # ── Header info rows ─────────────────────────────────────────────────────
        dur_text = batch.duration_display
        if start_date and end_date:
            dates_text = (
                f"START DATE: {start_date.strftime('%d %b %Y').upper()}   "
                f"END DATE: {end_date.strftime('%d %b %Y').upper()}"
            )
        else:
            dates_text = ""

        row_h1 = [para(f"<b>COLLEGE NAME :</b>  {batch.name}", size=8, align='LEFT')]
        row_h1 += [''] * (n_dates)
        row_h1 += [para(f"<b>INTERNSHIP :</b>  {batch.course}", size=8, align='LEFT')]
        row_h1 += [''] * (n_dates - 1)
        page_data.append(row_h1[:title_span])

        row_h2 = [para(f"<b>DURATION :</b>  {dur_text}", size=8, align='LEFT')]
        row_h2 += [''] * (n_dates)
        row_h2 += [para(dates_text, size=8, align='LEFT')]
        row_h2 += [''] * (n_dates - 1)
        page_data.append(row_h2[:title_span])

        # ── Date header row ───────────────────────────────────────────────────────
        row_dates = [para('', size=8)]
        for d in dates:
            row_dates.append(para(d.strftime('%d/%m/%Y'), bold=True, size=8))
            row_dates.append('')
        page_data.append(row_dates)

        # ── Sub-header row ────────────────────────────────────────────────────────
        row_sub = [para('NAME', bold=True, size=8)]
        for _ in dates:
            row_sub.append(para('TIME', bold=True, size=8))
            row_sub.append(para('Attendance', bold=True, size=8))
        page_data.append(row_sub)

        # ── Student data rows ─────────────────────────────────────────────────────
        for student in students:
            srec = lookup.get(student.id, {})
            row  = [para(student.name, size=8, color=C_RED, align='LEFT')]
            for d in dates:
                rec = srec.get(d)
                if rec:
                    if rec.login_time and rec.logout_time:
                        tv = (f"{rec.login_time.strftime('%I:%M%p').lstrip('0')}"
                              f"-{rec.logout_time.strftime('%I:%M%p').lstrip('0')}")
                    elif rec.login_time:
                        tv = rec.login_time.strftime('%I:%M%p').lstrip('0')
                    else:
                        tv = ''
                    av = rec.get_status_display()
                else:
                    tv = av = ''
                row.append(para(tv, size=7))
                row.append(para(av, size=7, bold=bool(av)))
            page_data.append(row)

        # Empty filler rows
        for _ in range(4):
            page_data.append([''] + [''] * (n_dates * 2))

        # ── Build table ───────────────────────────────────────────────────────────
        tbl = Table(page_data, colWidths=col_widths, repeatRows=0)

        # Merge spans: title, header rows, date cells
        n_stu = len(students)
        span_cmds = [
            # Title spans all columns
            ('SPAN', (0, 0), (title_span - 1, 0)),
            # Header info: each half spans half
            ('SPAN', (0, 1), (n_dates, 1)),
            ('SPAN', (n_dates + 1, 1), (title_span - 1, 1)),
            ('SPAN', (0, 2), (n_dates, 2)),
            ('SPAN', (n_dates + 1, 2), (title_span - 1, 2)),
            # Date merged pairs
        ]
        for i in range(n_dates):
            col = 1 + i * 2
            span_cmds.append(('SPAN', (col, 3), (col + 1, 3)))

        style_cmds = span_cmds + [
            # Outer grid
            ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#AAAAAA')),
            # Title
            ('BACKGROUND', (0, 0), (-1, 0), C_HDR),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, 0), 13),
            ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
            # Header info rows
            ('BACKGROUND', (0, 1), (-1, 2), C_HDR),
            ('FONTSIZE',   (0, 1), (-1, 2), 8),
            ('ALIGN',      (0, 1), (-1, 2), 'LEFT'),
            # Date row
            ('BACKGROUND', (0, 3), (-1, 3), C_DATE),
            ('FONTNAME',   (0, 3), (-1, 3), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 3), (-1, 3), 8),
            ('ALIGN',      (0, 3), (-1, 3), 'CENTER'),
            # Sub-header
            ('BACKGROUND', (0, 4), (-1, 4), C_SUB),
            ('FONTNAME',   (0, 4), (-1, 4), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 4), (-1, 4), 8),
            ('ALIGN',      (0, 4), (-1, 4), 'CENTER'),
            # Data rows
            ('FONTSIZE',   (0, 5), (-1, -1), 7),
            ('ALIGN',      (0, 5), (-1, -1), 'CENTER'),
            ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 5), (-1, 4 + n_stu),
             [colors.white, colors.HexColor('#F5F5F7')]),
            # Row heights
            ('ROWHEIGHT',  (0, 0), (0, 0), 24),
            ('ROWHEIGHT',  (0, 1), (0, 2), 15),
            ('ROWHEIGHT',  (0, 3), (0, 4), 16),
        ]

        # Colour attendance cells
        for s_i, student in enumerate(students):
            srec = lookup.get(student.id, {})
            data_row = 5 + s_i
            for d_i, d in enumerate(dates):
                rec = srec.get(d)
                if rec:
                    col_a = 2 + d_i * 2
                    fill  = (C_PRES if rec.status == 'present'
                             else C_ABS if rec.status == 'absent'
                             else C_NC)
                    style_cmds.append(('BACKGROUND', (col_a, data_row), (col_a, data_row), fill))

        tbl.setStyle(TableStyle(style_cmds))
        story.append(tbl)

        # ── Footer (last chunk only) ──────────────────────────────────────────────
        if ch_idx == len(chunks) - 1:
            story.append(Spacer(1, 8*mm))
            story.append(HRFlowable(width='100%', thickness=0.5,
                                    color=colors.HexColor('#AAAAAA')))
            story.append(Spacer(1, 4*mm))

            notes = [
                "1. This attendance sheet is prepared based on daily attendance records maintained by the organization.",
                "2. The details mentioned above are true and correct to the best of our knowledge.",
                "3. This document is issued only for submission to the concerned college/university.",
                "4. Attendance has been verified and approved by the undersigned authority.",
            ]
            footer_data = [
                [Paragraph('<b>Notes</b>', styles['Normal']),
                 Paragraph('<b>Verification</b>', styles['Normal'])],
                ['', ''],
            ]
            for i, note in enumerate(notes):
                left_cell = Paragraph(note, note_style)
                right_cell = ''
                if i == 1:
                    right_cell = Paragraph(
                        'Signature of Mentor/Supervisor: _______________________', ver_style)
                elif i == 3:
                    right_cell = Paragraph(
                        'Name &amp; Designation: _______________ / _______________', ver_style)
                footer_data.append([left_cell, right_cell])

            half = sum(col_widths) / 2
            ft = Table(footer_data, colWidths=[half, half])
            ft.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
            ]))
            story.append(ft)

        if ch_idx < len(chunks) - 1:
            from reportlab.platypus import PageBreak
            story.append(PageBreak())

    doc.build(story)
    buf.seek(0)

    period_slug = period_label.replace(' ', '_')
    fname = f"Batch_{batch.name.replace(' ', '_')}_{period_slug}_Attendance.pdf"
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response
